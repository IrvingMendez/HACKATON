import openpyxl

excel= openpyxl.load_workbook("ejemplo.xlsx")
 
gabriel=excel.active

for row in range (0, gabriel.max_row):
    _row=[row,]

    for col in gabriel.iter_cols(0,gabriel.max_column):
        print(col[row].value)

valor=gabriel["B1"].value

if valor == 1:
    consumido= gabriel["C1"].value

print (consumido, " watts consumidos")