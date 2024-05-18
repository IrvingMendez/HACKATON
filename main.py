import openpyxl
import os
import matlab.engine

excel = openpyxl.load_workbook("BaseDeDatos.xlsx")
libro = excel.active

selected_articles = []

def mostrarAreas():
    os.system("cls")
    print("¿De qué area es su articulo?")
    columna = libro["B"]

    for i, celda in enumerate(columna, start=1):
        if celda.value is not None:
            print(i, celda.value)

    opcion = int(input("Ingrese una opción: "))

    for celda in columna:
        if celda.row == opcion:
            area = celda.value
            break
    else:
        print("Opción inválida")
        return

    mostrarArticulos(area)

def mostrarArticulos(area):
    os.system("cls")
    print("¿Qué marca de artículo es?")
    fila = libro["C"]
    
    if area == "Dispositivo: Lavadora":
        rango = fila[:5]
    elif area == "Dispositivo: Televisores":
        rango = fila[5:10]
    elif area == "Dispositivo: Sistema de sonido":
        rango = fila[10:15]
    elif area == "Dispositivo: Ventilador":
        rango = fila[15:20]
    elif area == "Dispositivo: Aire acondicionado":
        rango = fila[20:25]
    elif area == "Dispositivo: Aspiradora":
        rango = fila[25:30]
    elif area == "Dispositivo: Plancha":
        rango = fila[30:35]
    elif area == "Dispositivo: Refrigerador":
        rango = fila[35:40]
    else:
        print("Área invalida")
        return

    for i, celda in enumerate(rango, start=rango[0].row):
        print(i, celda.value)

    opcion = int(input("Ingrese una opcion: "))
    os.system("cls")
    fila = libro[opcion]

    articulo = []
    for celda in fila:
        if celda.value is not None:
            print(celda.value)
            articulo.append(celda.value)

    opcion1 = int(input("¿Los datos son correctos? \n1. Sí \n2. No \nIngrese una opcion: "))
    if opcion1 == 1:
        if len(selected_articles) < 30:
            selected_articles.append(opcion)  
            print("Articulo guardado exitosamente")
        else:
            print("No se pueden guardar más de 15 articulos")
    elif opcion1 == 2:
        print("VOLVIENDO A LA PANTALLA ANTERIOR")
        mostrarAreas()

def mostrarArticulosGuardados():
    os.system("cls")
    if not selected_articles:
        print("No hay artículos guardados.")
    else:
        print("Artículos guardados:")
        for i, articulo in enumerate(selected_articles, start=1):
            fila = libro[articulo]
            print(f"Artículo {i}:")
            for celda in fila:
                if celda.value is not None:
                    print(celda.value)
            print("-----------------")
    watts_consumidos = WattsConsumidos()
    print("WATTS CONSUMIDOS:", watts_consumidos)
    if watts_consumidos <2000:
        print ("SIGUE ASI ESTAS CONSUMIENDO POCA ENERGIA")
    elif watts_consumidos>=2000 and watts_consumidos<7000:
        print("TU CONSUMO DE ENERGIA ES MODERADO")
    elif watts_consumidos>=7000 and watts_consumidos<10000:
        print("PRECAUSION TU CONSUMO ES ALTO PODRIAS TENER UNA SOBRECARGA")
    elif watts_consumidos>10000:
        print("PELIGRO DE SOBRECARGA")

    input("Presione Enter para volver al menú...")

def eliminarArticulo():
    os.system("cls")
    if not selected_articles:
        print("No hay artículos guardados.")
    else:
        print("Seleccione el artículo que desea eliminar:")
        for i, articulo in enumerate(selected_articles, start=1):
            fila = libro[articulo]
            print(f"{i}. {fila[0].value}") 
        
        opcion = int(input("Ingrese el número del artículo que desea eliminar: "))
        if 1 <= opcion <= len(selected_articles):
            deleted_article = selected_articles.pop(opcion - 1)
            print(f"Artículo '{libro[deleted_article][0].value}' eliminado exitosamente.")
        else:
            print("Opción inválida.")
    input("Presione Enter para volver al menu...")

def WattsConsumidos():
    valores = obtenerValores()
    if not valores:
        return 0


    eng = matlab.engine.start_matlab()
    matlab_valores = matlab.double(valores, is_complex=False)
    resultado = eng.GraficaSuma(matlab_valores)

    eng.plot_data(matlab_valores, nargout=0)
    input ("Presiona una tecla para salir de la grafica")
    eng.quit()
    return resultado



def obtenerValores():
    valores = []
    for fila_num in selected_articles:
        celda_f = libro.cell(row=fila_num, column=6)
        if celda_f.value is not None:
            if celda_f.value == 1:
                celda_valor = libro.cell(row=fila_num, column=5).value
            else:
                celda_valor = libro.cell(row=fila_num, column=4).value
            
            if celda_valor is not None and isinstance(celda_valor, (int, float)):
                valores.append(celda_valor)
            else:
                print(f"Valor inválido en la fila {fila_num}: {celda_valor}")  
    return valores

def menu():
    while True:
        os.system("cls")
        print("""MENU:
        1. Agregar un artículo
        2. Eliminar artículo
        3. Mostrar artículos guardados
        4. Salir""")
        opcion = int(input("Ingrese una opción: "))

        if opcion == 1:
            os.system("cls")
            mostrarAreas()
        elif opcion == 2:
            eliminarArticulo()
        elif opcion == 3:
            mostrarArticulosGuardados()
        elif opcion == 4:
            break
        else:
            print("Opción inválida")
            input("Presione Enter para volver al menú...")

menu()